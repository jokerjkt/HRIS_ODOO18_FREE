# -*- coding: utf-8 -*-
"""
hr.attendance.connector.csv — Generic CSV/Excel Connector
===========================================================
Connector universal untuk import file CSV/Excel dari berbagai merek mesin:
- ZKTeco (export via software atau USB)
- Solution
- Fingerspot
- ATT2000
- Generic/Universal

Format yang didukung:
- CSV (comma atau tab separated)
- Excel (.xlsx)
- DAT (tab separated)
"""
import csv
import io
import os
from datetime import datetime

from odoo import models, api
from odoo.exceptions import ValidationError

try:
    import openpyxl
except ImportError:
    openpyxl = None


# Column name mappings — semua kemungkinan nama kolom dari berbagai merek
COLUMN_MAPS = {
    'user_id': [
        'user id', 'userid', 'user_id', 'empcode', 'emp code', 'employee id',
        'employee_id', 'pin', 'pin Code', 'no', 'code', 'employee code',
        'nik', 'nomor', 'id', 'finger id',
    ],
    'datetime': [
        'datetime', 'date/time', 'date time', 'date', 'time', 'timestamp',
        'waktu', 'tanggal', 'tgl', 'check time', 'atttime', 'att time',
        'attendance time', 'scan time',
    ],
    'status': [
        'status', 'check type', 'check_type', 'type', 'in/out', 'inout',
        'checkin/checkout', 'tipe', 'jenis', 'punch', 'punch type',
        'attstatus', 'att status',
    ],
    'verify': [
        'verify', 'verified', 'verifymode', 'verify mode', 'verification',
        'method', 'metode', 'verifikasi', 'auth', 'authentication',
        'fingerprint', 'fp', 'face', 'card',
    ],
    'workcode': [
        'workcode', 'work code', 'work_code', 'wc',
    ],
}

# Status mappings
STATUS_IN = ['0', 'i', 'in', 'check in', 'checkin', 'ci', 'masuk', 'hadir']
STATUS_OUT = ['1', 'o', 'out', 'check out', 'checkout', 'co', 'keluar', 'pulang']

# Verify mode mappings
VERIFY_MAP = {
    'password': '0', 'pw': '0', 'pwd': '0', 'pass': '0', 'password': '0',
    'fingerprint': '1', 'fp': '1', 'finger': '1', 'sidik jari': '1',
    'card': '2', 'rfid': '2', 'kartu': '2', 'badge': '2', 'icard': '2',
    'face': '3', 'wajah': '3', 'facial': '3',
    'iris': '4', 'mata': '4',
}

# Date formats to try
DATE_FORMATS = [
    '%Y-%m-%d %H:%M:%S',
    '%Y-%m-%d %H:%M',
    '%Y/%m/%d %H:%M:%S',
    '%Y/%m/%d %H:%M',
    '%d-%m-%Y %H:%M:%S',
    '%d-%m-%Y %H:%M',
    '%d/%m/%Y %H:%M:%S',
    '%d/%m/%Y %H:%M',
    '%m/%d/%Y %H:%M:%S',
    '%m/%d/%Y %H:%M',
    '%d %b %Y %H:%M:%S',
    '%d %b %Y %H:%M',
    '%Y%m%d%H%M%S',
    '%Y%m%d%H%M',
]


class HrAttendanceConnectorCsv(models.Model):
    _name = 'hr.attendance.connector.csv'
    _description = 'CSV/Excel Connector'
    _inherit = 'hr.attendance.connector'

    def test_connection(self, device):
        """CSV import tidak perlu test koneksi."""
        return {'success': True, 'message': 'File import mode'}

    def parse_file(self, file_content, file_name='data.csv'):
        """Parse file CSV/Excel dan return list of dicts."""
        ext = os.path.splitext(file_name)[1].lower()

        if ext in ('.xlsx', '.xls'):
            return self._parse_excel(file_content)
        else:
            return self._parse_csv(file_content, file_name)

    def _parse_csv(self, file_content, file_name='data.csv'):
        """Parse CSV/DAT file."""
        # Decode bytes jika perlu
        if isinstance(file_content, bytes):
            # Coba beberapa encoding
            for enc in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    file_content = file_content.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue

        # Detect separator
        first_line = file_content.split('\n')[0]
        if '\t' in first_line:
            separator = '\t'
        elif ';' in first_line:
            separator = ';'
        else:
            separator = ','

        reader = csv.DictReader(io.StringIO(file_content), delimiter=separator)
        return self._process_rows(reader)

    def _parse_excel(self, file_content):
        """Parse Excel file."""
        if not openpyxl:
            raise ValidationError(
                'Library openpyxl tidak terinstall. '
                'Install dengan: pip install openpyxl'
            )

        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
        ws = wb.active

        # Get headers
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value or '').strip())

        # Get data rows
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, cell_value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(cell_value or '').strip()
            if any(row_dict.values()):
                rows.append(row_dict)

        return self._process_rows(rows)

    def _process_rows(self, rows):
        """Process rows (dict reader or list of dicts) dan return parsed data."""
        results = []

        for row in rows:
            # Normalize column names
            normalized = {}
            for key, value in row.items():
                if key:
                    normalized[key.strip().lower()] = value

            # Find column mappings
            user_id_col = self._find_column(normalized, 'user_id')
            datetime_col = self._find_column(normalized, 'datetime')
            status_col = self._find_column(normalized, 'status')
            verify_col = self._find_column(normalized, 'verify')

            if not user_id_col or not datetime_col:
                continue

            # Parse values
            user_id = str(row.get(user_id_col, '')).strip()
            if not user_id:
                continue

            dt_str = str(row.get(datetime_col, '')).strip()
            timestamp = self._parse_datetime(dt_str)
            if not timestamp:
                continue

            # Status
            status_str = str(row.get(status_col, '0')).strip().lower() if status_col else '0'
            punch_type = self._map_status(status_str)

            # Verify mode
            verify_str = str(row.get(verify_col, '1')).strip().lower() if verify_col else '1'
            verify_mode = self._map_verify(verify_str)

            # Raw data
            raw = '\t'.join(f'{k}={v}' for k, v in row.items() if k)

            results.append({
                'employee_pin': user_id,
                'timestamp': timestamp,
                'punch_type': punch_type,
                'verify_mode': verify_mode,
                'raw_data': raw,
            })

        return results

    def _find_column(self, normalized_row, field_type):
        """Cari kolom yang cocok berdasarkan tipe field."""
        candidates = COLUMN_MAPS.get(field_type, [])
        for candidate in candidates:
            if candidate in normalized_row:
                return candidate
        # Return original key (case-sensitive) if found
        for key in normalized_row:
            if key.lower() in candidates:
                return key
        return None

    def _parse_datetime(self, dt_str):
        """Parse datetime string ke Python datetime."""
        if not dt_str:
            return None

        dt_str = dt_str.strip()

        # Coba semua format
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(dt_str, fmt)
            except ValueError:
                continue

        # Coba parse sebagai timestamp (unix)
        try:
            ts = float(dt_str)
            if ts > 1e9:  # Unix timestamp dalam detik
                return datetime.fromtimestamp(ts)
        except (ValueError, TypeError, OSError):
            pass

        return None

    def _map_status(self, status_str):
        """Map status string ke punch_type."""
        status_str = status_str.strip().lower()
        if status_str in STATUS_IN:
            return '0'
        elif status_str in STATUS_OUT:
            return '1'
        # Default: jika angka, gunakan langsung
        if status_str in ('0', '1', '2', '3', '4', '5'):
            return status_str
        return '0'  # Default check in

    def _map_verify(self, verify_str):
        """Map verify string ke verify_mode code."""
        verify_str = verify_str.strip().lower()
        # Jika sudah angka
        if verify_str in ('0', '1', '2', '3', '4'):
            return verify_str
        # Cari di mapping
        for key, val in VERIFY_MAP.items():
            if key in verify_str:
                return val
        return '1'  # Default fingerprint

    def action_parse_file(self, device, file_content, file_name):
        """Parse file dan buat device logs."""
        logs_data = self.parse_file(file_content, file_name)
        if not logs_data:
            raise ValidationError(
                'Tidak ada data yang dapat diparsing dari file ini.\n'
                'Pastikan file memiliki kolom: User ID, DateTime, Status'
            )
        return self._create_device_logs(device, logs_data)
